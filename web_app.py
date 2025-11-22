# -*- coding: utf-8 -*-
"""
학원 등원/하원 웹 앱
PC, 모바일, 태블릿에서 모두 사용 가능
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for
import openpyxl
import os
import json
from datetime import datetime
from sms_sender import send_sms

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False

# 설정 파일 경로
EXCEL_FILE = os.getenv('EXCEL_FILE', '202511_자동알림.xlsx')
CONFIG_FILE = os.getenv('CONFIG_FILE', 'config.json')
SMS_CONFIG_FILE = os.getenv('SMS_CONFIG_FILE', 'sms_config.json')

def init_excel_file():
    """Excel 파일이 없으면 생성"""
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel 파일이 없습니다. 새로 생성합니다: {EXCEL_FILE}")
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 헤더 생성
        ws['A1'] = '이름'
        ws['B1'] = '연락처'
        ws['C1'] = '상태'
        ws['D1'] = '납입일'
        
        # 샘플 데이터 (선택사항)
        ws['A2'] = '홍길동'
        ws['B2'] = '01012345678'
        ws['C2'] = 0
        ws['D2'] = ''
        
        wb.save(EXCEL_FILE)
        print(f"Excel 파일 생성 완료: {EXCEL_FILE}")

def load_config():
    """설정 파일 로드 (환경 변수 우선)"""
    # 환경 변수에서 먼저 읽기 (Render, Railway 등 호스팅 서비스용)
    if os.getenv('ACADEMY_NAME'):
        return {
            "academy_name": os.getenv('ACADEMY_NAME', 'OO학원'),
            "name_column": os.getenv('NAME_COLUMN', 'A'),
            "phone_column": os.getenv('PHONE_COLUMN', 'B'),
            "status_column": os.getenv('STATUS_COLUMN', 'C'),
            "payment_column": os.getenv('PAYMENT_COLUMN', 'D'),
            "start_row": int(os.getenv('START_ROW', '2'))
        }
    
    # 로컬 환경에서는 config.json 사용
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    # 기본값
    return {
        "academy_name": "OO학원",
        "name_column": "A",
        "phone_column": "B",
        "status_column": "C",
        "payment_column": "D",
        "start_row": 2
    }

def read_students():
    """학생 목록 읽기"""
    config = load_config()
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        students = []
        row = config['start_row']
        
        while True:
            name_cell = f"{config['name_column']}{row}"
            phone_cell = f"{config['phone_column']}{row}"
            status_cell = f"{config['status_column']}{row}"
            payment_cell = f"{config['payment_column']}{row}"
            
            name = ws[name_cell].value
            
            if not name:
                break
            
            phone = ws[phone_cell].value
            status = ws[status_cell].value
            payment_date = ws[payment_cell].value
            
            if status is None:
                status = 0
            else:
                try:
                    status = int(status)
                except:
                    status = 0
            
            students.append({
                'row': row,
                'name': name,
                'phone': str(phone) if phone else '',
                'status': status,
                'payment_date': payment_date,
                'payment_status': '납입완료' if payment_date else '미납'
            })
            
            row += 1
        
        wb.close()
        return students
        
    except Exception as e:
        print(f"엑셀 파일 읽기 오류: {e}")
        return []

def update_status(row, new_status):
    """학생 상태 업데이트"""
    config = load_config()
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        status_cell = f"{config['status_column']}{row}"
        ws[status_cell].value = new_status
        
        wb.save(EXCEL_FILE)
        wb.close()
        return True
        
    except Exception as e:
        print(f"상태 업데이트 오류: {e}")
        return False

# 앱 시작 시 Excel 파일 초기화 (Gunicorn 실행 시에도 작동)
init_excel_file()

@app.route('/')
def index():
    """메인 페이지"""
    config = load_config()
    students = read_students()
    return render_template('index.html', 
                         students=students, 
                         academy_name=config.get('academy_name', 'OO학원'))

@app.route('/api/students')
def get_students():
    """학생 목록 API"""
    students = read_students()
    return jsonify(students)

@app.route('/api/checkin/<int:row>', methods=['POST'])
def checkin(row):
    """등원 처리 API"""
    config = load_config()
    students = read_students()
    
    # 해당 학생 찾기
    student = next((s for s in students if s['row'] == row), None)
    
    if not student:
        return jsonify({'success': False, 'message': '학생을 찾을 수 없습니다.'}), 404
    
    # 이미 등원중인지 확인
    if student['status'] == 1:
        return jsonify({'success': False, 'message': f"{student['name']}님은 이미 등원중입니다."})
    
    # 상태 업데이트
    if update_status(row, 1):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        message = f'"{student["name"]}"님이 "{config.get("academy_name", "OO학원")}"에 등원하였습니다.'
        
        # 메시지 전송
        if student['phone']:
            send_sms(student['phone'], message, student_name=student['name'])
        
        return jsonify({
            'success': True, 
            'message': f"{student['name']}님 등원 처리 완료",
            'timestamp': timestamp,
            'notification': message
        })
    else:
        return jsonify({'success': False, 'message': '상태 업데이트 실패'}), 500

@app.route('/api/checkout/<int:row>', methods=['POST'])
def checkout(row):
    """하원 처리 API"""
    config = load_config()
    students = read_students()
    
    # 해당 학생 찾기
    student = next((s for s in students if s['row'] == row), None)
    
    if not student:
        return jsonify({'success': False, 'message': '학생을 찾을 수 없습니다.'}), 404
    
    # 이미 하원 상태인지 확인
    if student['status'] == 0:
        return jsonify({'success': False, 'message': f"{student['name']}님은 이미 하원 상태입니다."})
    
    # 상태 업데이트
    if update_status(row, 0):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        message = f'"{student["name"]}"님이 "{config.get("academy_name", "OO학원")}"에서 하원하였습니다.'
        
        # 메시지 전송
        if student['phone']:
            send_sms(student['phone'], message, student_name=student['name'])
        
        return jsonify({
            'success': True, 
            'message': f"{student['name']}님 하원 처리 완료",
            'timestamp': timestamp,
            'notification': message
        })
    else:
        return jsonify({'success': False, 'message': '상태 업데이트 실패'}), 500

@app.route('/api/payment/<int:row>', methods=['POST'])
def register_payment(row):
    """원비 납입 등록 API"""
    config = load_config()
    students = read_students()
    
    # 해당 학생 찾기
    student = next((s for s in students if s['row'] == row), None)
    
    if not student:
        return jsonify({'success': False, 'message': '학생을 찾을 수 없습니다.'}), 404
    
    # 납입일 가져오기
    data = request.get_json()
    payment_date = data.get('payment_date')
    
    # 납입 정보 업데이트
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        payment_cell = f"{config['payment_column']}{row}"
        ws[payment_cell].value = payment_date
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        if payment_date:
            return jsonify({
                'success': True,
                'message': f"{student['name']}님 원비 납입 등록 완료",
                'payment_date': payment_date
            })
        else:
            return jsonify({
                'success': True,
                'message': f"{student['name']}님 납입 정보 삭제 완료"
            })
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'업데이트 오류: {e}'}), 500

@app.route('/api/send_message/<int:row>', methods=['POST'])
def send_message(row):
    """메시지 수동 발송 API"""
    config = load_config()
    students = read_students()
    
    # 해당 학생 찾기
    student = next((s for s in students if s['row'] == row), None)
    
    if not student:
        return jsonify({'success': False, 'message': '학생을 찾을 수 없습니다.'}), 404
    
    if not student['phone']:
        return jsonify({'success': False, 'message': '연락처가 없습니다.'}), 400
    
    # 메시지 타입 가져오기
    data = request.get_json()
    msg_type = data.get('type')  # 'checkin', 'checkout', 'payment_request'
    custom_message = data.get('message', '')
    
    academy_name = config.get('academy_name', 'OO학원')
    
    # 메시지 생성
    if msg_type == 'checkin':
        message = f'"{student["name"]}"님이 "{academy_name}"에 등원하였습니다.'
    elif msg_type == 'checkout':
        message = f'"{student["name"]}"님이 "{academy_name}"에서 하원하였습니다.'
    elif msg_type == 'payment_request':
        if custom_message:
            message = custom_message
        else:
            message = f'안녕하세요, {academy_name}입니다.\n{student["name"]}님의 이번 달 원비 납입을 부탁드립니다.'
    else:
        return jsonify({'success': False, 'message': '잘못된 메시지 타입입니다.'}), 400
    
    # 메시지 발송
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    send_sms(student['phone'], message, student_name=student['name'])
    
    return jsonify({
        'success': True,
        'message': f"{student['name']}님에게 메시지 발송 완료",
        'timestamp': timestamp
    })

@app.route('/api/edit_phone/<int:row>', methods=['POST'])
def edit_phone(row):
    """연락처 수정 API"""
    config = load_config()
    students = read_students()
    
    # 해당 학생 찾기
    student = next((s for s in students if s['row'] == row), None)
    
    if not student:
        return jsonify({'success': False, 'message': '학생을 찾을 수 없습니다.'}), 404
    
    # 새 연락처 가져오기
    data = request.get_json()
    new_phone = data.get('phone')
    
    if not new_phone or not new_phone.strip():
        return jsonify({'success': False, 'message': '연락처를 입력해주세요.'}), 400
    
    # 연락처 업데이트
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        phone_cell = f"{config['phone_column']}{row}"
        ws[phone_cell].value = new_phone.strip()
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        return jsonify({
            'success': True,
            'message': f"{student['name']}님 연락처 수정 완료",
            'new_phone': new_phone.strip()
        })
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'수정 오류: {e}'}), 500

@app.route('/api/add_student', methods=['POST'])
def add_student():
    """학생 등록 API"""
    config = load_config()
    
    # 데이터 가져오기
    data = request.get_json()
    name = data.get('name')
    phone = data.get('phone')
    payment_date = data.get('payment_date', '')
    
    if not name or not name.strip():
        return jsonify({'success': False, 'message': '이름을 입력해주세요.'}), 400
    
    if not phone or not phone.strip():
        return jsonify({'success': False, 'message': '연락처를 입력해주세요.'}), 400
    
    # 엑셀에 추가
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # 마지막 행 찾기
        last_row = config['start_row']
        while ws[f"{config['name_column']}{last_row}"].value:
            last_row += 1
        
        # 새 학생 정보 입력
        ws[f"{config['name_column']}{last_row}"].value = name.strip()
        ws[f"{config['phone_column']}{last_row}"].value = phone.strip()
        ws[f"{config['status_column']}{last_row}"].value = 0  # 하원 상태
        if payment_date and payment_date.strip():
            ws[f"{config['payment_column']}{last_row}"].value = payment_date.strip()
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        return jsonify({
            'success': True,
            'message': f"{name}님 등록 완료",
            'student': {
                'name': name.strip(),
                'phone': phone.strip(),
                'payment_date': payment_date.strip() if payment_date else None
            }
        })
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'등록 오류: {e}'}), 500

@app.route('/api/delete_student/<int:row>', methods=['DELETE'])
def delete_student(row):
    """학생 삭제 API"""
    config = load_config()
    students = read_students()
    
    # 해당 학생 찾기
    student = next((s for s in students if s['row'] == row), None)
    
    if not student:
        return jsonify({'success': False, 'message': '학생을 찾을 수 없습니다.'}), 404
    
    # 엑셀에서 삭제
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # 해당 행 삭제
        ws.delete_rows(row, 1)
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        return jsonify({
            'success': True,
            'message': f"{student['name']}님 삭제 완료"
        })
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'삭제 오류: {e}'}), 500

@app.route('/mobile')
def mobile():
    """모바일 최적화 페이지"""
    config = load_config()
    students = read_students()
    return render_template('mobile.html', 
                         students=students, 
                         academy_name=config.get('academy_name', 'OO학원'))

if __name__ == '__main__':
    # Excel 파일 초기화 (없으면 생성)
    init_excel_file()
    
    # Render 등 호스팅 서비스는 PORT 환경 변수를 제공
    port = int(os.getenv('PORT', 5000))
    
    # 0.0.0.0으로 설정하면 같은 네트워크의 다른 기기에서 접근 가능
    print()
    print("=" * 60)
    print("학원 등원/하원 웹 앱 시작")
    print("=" * 60)
    print()
    print(f"PC에서 접속: http://localhost:{port}")
    print(f"모바일에서 접속: http://[PC의_IP주소]:{port}")
    print()
    print("예시:")
    print(f"  - PC: http://localhost:{port}")
    print(f"  - 모바일: http://192.168.0.10:{port}")
    print()
    print("종료: Ctrl+C")
    print("=" * 60)
    print()
    
    app.run(host='0.0.0.0', port=port, debug=False)


