# -*- coding: utf-8 -*-
"""
학원 등원/하원 입력 프로그램
엑셀 파일을 직접 수정하지 않고 프로그램에서 상태를 변경
"""

import openpyxl
import os
import json
from datetime import datetime

class AttendanceInput:
    def __init__(self, excel_file='202511_자동알림.xlsx', config_file='config.json'):
        self.excel_file = excel_file
        self.config = self.load_config(config_file)
        
    def load_config(self, config_file):
        """설정 파일 로드"""
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            return {
                "name_column": "A",
                "phone_column": "B",
                "status_column": "C",
                "start_row": 2
            }
    
    def read_students(self):
        """학생 목록 읽기"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            students = []
            row = self.config['start_row']
            
            while True:
                name_cell = f"{self.config['name_column']}{row}"
                phone_cell = f"{self.config['phone_column']}{row}"
                status_cell = f"{self.config['status_column']}{row}"
                payment_cell = f"{self.config['payment_column']}{row}"
                
                name = ws[name_cell].value
                
                if not name:
                    break
                
                phone = ws[phone_cell].value
                status = ws[status_cell].value
                payment_date = ws[payment_cell].value
                
                if status is None:
                    status = 0
                else:
                    try:
                        status = int(status)
                    except:
                        status = 0
                
                students.append({
                    'row': row,
                    'name': name,
                    'phone': str(phone) if phone else '',
                    'status': status,
                    'payment_date': payment_date
                })
                
                row += 1
            
            wb.close()
            return students
            
        except Exception as e:
            print(f"엑셀 파일 읽기 오류: {e}")
            return []
    
    def update_status(self, row, new_status):
        """학생 상태 업데이트"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            status_cell = f"{self.config['status_column']}{row}"
            ws[status_cell].value = new_status
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            print(f"상태 업데이트 오류: {e}")
            return False
    
    def display_students(self, students):
        """학생 목록 표시"""
        print()
        print("=" * 105)
        print(f"{'번호':<6}{'이름':<12}{'연락처':<18}{'현재 상태':<15}{'원비 납입':<30}")
        print("=" * 105)
        
        for idx, student in enumerate(students, 1):
            status_text = "✓ 등원중" if student['status'] == 1 else "○ 하원"
            payment_date = student.get('payment_date')
            if payment_date:
                payment_text = f"납입완료 ({payment_date})"
            else:
                payment_text = "미납"
            print(f"{idx:<6}{student['name']:<12}{student['phone']:<18}{status_text:<15}{payment_text:<30}")
        
        print("=" * 105)
        print()
    
    def run(self):
        """메인 실행 루프"""
        print()
        print("█" * 70)
        print("█" + " " * 68 + "█")
        print("█" + " " * 20 + "학원 등원/하원 입력 시스템" + " " * 22 + "█")
        print("█" + " " * 68 + "█")
        print("█" * 70)
        print()
        
        if not os.path.exists(self.excel_file):
            print(f"오류: 엑셀 파일을 찾을 수 없습니다: {self.excel_file}")
            input("\nEnter 키를 눌러 종료...")
            return
        
        while True:
            students = self.read_students()
            
            if not students:
                print("등록된 학생이 없습니다.")
                input("\nEnter 키를 눌러 종료...")
                break
            
            self.display_students(students)
            
            print("명령어:")
            print("  - 숫자 입력: 해당 학생의 상태 변경")
            print("  - 'q' 또는 'exit': 종료")
            print("  - 'r' 또는 'refresh': 새로고침")
            print()
            
            choice = input("선택하세요: ").strip().lower()
            
            if choice in ['q', 'exit', '종료']:
                print("\n프로그램을 종료합니다.")
                break
            
            if choice in ['r', 'refresh', '새로고침']:
                print("\n새로고침 중...")
                continue
            
            try:
                idx = int(choice)
                
                if 1 <= idx <= len(students):
                    student = students[idx - 1]
                    current_status = student['status']
                    
                    print()
                    print(f"선택한 학생: {student['name']}")
                    print(f"현재 상태: {'등원중' if current_status == 1 else '하원'}")
                    print()
                    print("1. 등원 처리")
                    print("2. 하원 처리")
                    print("0. 취소")
                    print()
                    
                    action = input("선택하세요: ").strip()
                    
                    if action == '1':
                        # 등원 처리
                        if current_status == 1:
                            print(f"\n⚠️  {student['name']}님은 이미 등원중입니다.")
                        else:
                            if self.update_status(student['row'], 1):
                                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                print(f"\n✓ [{timestamp}] {student['name']}님이 등원 처리되었습니다.")
                                print(f"  메시지: \"{student['name']}\"님이 \"{self.config.get('academy_name', 'OO학원')}\"에 등원하였습니다.")
                            else:
                                print(f"\n✗ 상태 변경에 실패했습니다.")
                    
                    elif action == '2':
                        # 하원 처리
                        if current_status == 0:
                            print(f"\n⚠️  {student['name']}님은 이미 하원 상태입니다.")
                        else:
                            if self.update_status(student['row'], 0):
                                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                print(f"\n✓ [{timestamp}] {student['name']}님이 하원 처리되었습니다.")
                                print(f"  메시지: \"{student['name']}\"님이 \"{self.config.get('academy_name', 'OO학원')}\"에서 하원하였습니다.")
                            else:
                                print(f"\n✗ 상태 변경에 실패했습니다.")
                    
                    elif action == '0':
                        print("\n취소되었습니다.")
                    
                    else:
                        print("\n잘못된 선택입니다.")
                    
                    input("\nEnter 키를 눌러 계속...")
                    
                else:
                    print(f"\n잘못된 번호입니다. 1-{len(students)} 사이의 숫자를 입력하세요.")
                    input("\nEnter 키를 눌러 계속...")
            
            except ValueError:
                print("\n잘못된 입력입니다. 숫자를 입력하세요.")
                input("\nEnter 키를 눌러 계속...")
            except Exception as e:
                print(f"\n오류 발생: {e}")
                input("\nEnter 키를 눌러 계속...")

if __name__ == "__main__":
    try:
        app = AttendanceInput()
        app.run()
    except KeyboardInterrupt:
        print("\n\n프로그램을 종료합니다.")
    except Exception as e:
        print(f"\n오류 발생: {e}")
        import traceback
        traceback.print_exc()
        input("\nEnter 키를 눌러 종료...")

