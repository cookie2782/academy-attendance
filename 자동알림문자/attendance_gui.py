# -*- coding: utf-8 -*-
"""
학원 등원/하원 입력 프로그램 (GUI 버전)
tkinter를 사용한 그래픽 인터페이스
"""

import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
import json
from datetime import datetime
import threading
import time

class AttendanceGUI:
    def __init__(self, excel_file='202511_자동알림.xlsx', config_file='config.json'):
        self.excel_file = excel_file
        self.config = self.load_config(config_file)
        self.root = tk.Tk()
        self.root.title("학원 등원/하원 입력 시스템")
        self.root.geometry("800x600")
        
        # 스타일 설정
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        self.setup_ui()
        self.load_students()
        
        # 자동 새로고침 (3초마다)
        self.auto_refresh()
    
    def load_config(self, config_file):
        """설정 파일 로드"""
        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
        else:
            config = {
                "academy_name": "OO학원",
                "name_column": "A",
                "phone_column": "B",
                "status_column": "C",
                "start_row": 2
            }
        return config
    
    def setup_ui(self):
        """UI 구성"""
        # 상단 프레임 (제목)
        top_frame = tk.Frame(self.root, bg="#2196F3", height=80)
        top_frame.pack(fill=tk.X, pady=0)
        
        title_label = tk.Label(
            top_frame, 
            text=f"{self.config.get('academy_name', 'OO학원')} 등원/하원 관리",
            font=("맑은 고딕", 20, "bold"),
            bg="#2196F3",
            fg="white"
        )
        title_label.pack(pady=20)
        
        # 중간 프레임 (학생 목록)
        middle_frame = tk.Frame(self.root, bg="white")
        middle_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # 트리뷰 (학생 목록 테이블)
        columns = ("번호", "이름", "연락처", "상태", "납입상태")
        self.tree = ttk.Treeview(middle_frame, columns=columns, show="headings", height=15)
        
        # 컬럼 설정
        self.tree.heading("번호", text="번호")
        self.tree.heading("이름", text="이름")
        self.tree.heading("연락처", text="연락처")
        self.tree.heading("상태", text="현재 상태")
        self.tree.heading("납입상태", text="원비 납입")
        
        self.tree.column("번호", width=60, anchor="center")
        self.tree.column("이름", width=120, anchor="center")
        self.tree.column("연락처", width=150, anchor="center")
        self.tree.column("상태", width=100, anchor="center")
        self.tree.column("납입상태", width=180, anchor="center")
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(middle_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 하단 프레임 (버튼)
        bottom_frame = tk.Frame(self.root, bg="white")
        bottom_frame.pack(fill=tk.X, padx=20, pady=20)
        
        # 버튼 스타일
        button_style = {
            "font": ("맑은 고딕", 12, "bold"),
            "width": 12,
            "height": 2
        }
        
        self.checkin_btn = tk.Button(
            bottom_frame,
            text="✓ 등원",
            bg="#4CAF50",
            fg="white",
            command=self.checkin,
            **button_style
        )
        self.checkin_btn.pack(side=tk.LEFT, padx=10)
        
        self.checkout_btn = tk.Button(
            bottom_frame,
            text="✗ 하원",
            bg="#FF5722",
            fg="white",
            command=self.checkout,
            **button_style
        )
        self.checkout_btn.pack(side=tk.LEFT, padx=10)
        
        self.refresh_btn = tk.Button(
            bottom_frame,
            text="🔄 새로고침",
            bg="#2196F3",
            fg="white",
            command=self.load_students,
            **button_style
        )
        self.refresh_btn.pack(side=tk.LEFT, padx=10)
        
        self.payment_btn = tk.Button(
            bottom_frame,
            text="💰 납입등록",
            bg="#FF9800",
            fg="white",
            command=self.register_payment,
            **button_style
        )
        self.payment_btn.pack(side=tk.LEFT, padx=10)
        
        # 구분선
        separator = tk.Frame(bottom_frame, width=2, bg="#ddd")
        separator.pack(side=tk.LEFT, fill=tk.Y, padx=20)
        
        # 문자 발송 버튼들
        msg_button_style = {
            "font": ("맑은 고딕", 11, "bold"),
            "width": 12,
            "height": 2
        }
        
        self.send_checkin_msg_btn = tk.Button(
            bottom_frame,
            text="📨 등원알림",
            bg="#9C27B0",
            fg="white",
            command=self.send_checkin_message,
            **msg_button_style
        )
        self.send_checkin_msg_btn.pack(side=tk.LEFT, padx=5)
        
        self.send_checkout_msg_btn = tk.Button(
            bottom_frame,
            text="📨 하원알림",
            bg="#673AB7",
            fg="white",
            command=self.send_checkout_message,
            **msg_button_style
        )
        self.send_checkout_msg_btn.pack(side=tk.LEFT, padx=5)
        
        self.send_payment_request_btn = tk.Button(
            bottom_frame,
            text="📨 납입요청",
            bg="#E91E63",
            fg="white",
            command=self.send_payment_request,
            **msg_button_style
        )
        self.send_payment_request_btn.pack(side=tk.LEFT, padx=5)
        
        # 구분선 2
        separator2 = tk.Frame(bottom_frame, width=2, bg="#ddd")
        separator2.pack(side=tk.LEFT, fill=tk.Y, padx=20)
        
        # 관리 버튼들
        manage_button_style = {
            "font": ("맑은 고딕", 11, "bold"),
            "width": 10,
            "height": 2
        }
        
        self.edit_phone_btn = tk.Button(
            bottom_frame,
            text="📞 연락처수정",
            bg="#00BCD4",
            fg="white",
            command=self.edit_phone,
            **manage_button_style
        )
        self.edit_phone_btn.pack(side=tk.LEFT, padx=5)
        
        self.add_student_btn = tk.Button(
            bottom_frame,
            text="➕ 학생등록",
            bg="#4CAF50",
            fg="white",
            command=self.add_student,
            **manage_button_style
        )
        self.add_student_btn.pack(side=tk.LEFT, padx=5)
        
        self.delete_student_btn = tk.Button(
            bottom_frame,
            text="🗑️ 학생삭제",
            bg="#F44336",
            fg="white",
            command=self.delete_student,
            **manage_button_style
        )
        self.delete_student_btn.pack(side=tk.LEFT, padx=5)
        
        # 상태 표시줄
        self.status_label = tk.Label(
            self.root,
            text=f"마지막 업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            font=("맑은 고딕", 9),
            bg="#f0f0f0",
            anchor="w"
        )
        self.status_label.pack(fill=tk.X, side=tk.BOTTOM)
    
    def read_students(self):
        """학생 목록 읽기"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            students = []
            row = self.config['start_row']
            
            while True:
                name_cell = f"{self.config['name_column']}{row}"
                phone_cell = f"{self.config['phone_column']}{row}"
                status_cell = f"{self.config['status_column']}{row}"
                payment_cell = f"{self.config['payment_column']}{row}"
                
                name = ws[name_cell].value
                
                if not name:
                    break
                
                phone = ws[phone_cell].value
                status = ws[status_cell].value
                payment_date = ws[payment_cell].value
                
                if status is None:
                    status = 0
                else:
                    try:
                        status = int(status)
                    except:
                        status = 0
                
                students.append({
                    'row': row,
                    'name': name,
                    'phone': str(phone) if phone else '',
                    'status': status,
                    'payment_date': payment_date
                })
                
                row += 1
            
            wb.close()
            return students
            
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일 읽기 오류: {e}")
            return []
    
    def load_students(self):
        """학생 목록 로드 및 표시"""
        # 기존 항목 삭제
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # 학생 목록 읽기
        students = self.read_students()
        
        # 트리뷰에 추가
        for idx, student in enumerate(students, 1):
            status_text = "✓ 등원중" if student['status'] == 1 else "○ 하원"
            
            # 납입 상태 확인
            payment_date = student.get('payment_date')
            if payment_date:
                payment_status = f"납입완료 ({payment_date})"
            else:
                payment_status = "미납"
            payment_tag = "paid" if payment_date else "unpaid"
            
            tag = "checkedin" if student['status'] == 1 else "checkedout"
            combined_tag = f"{tag}_{payment_tag}"
            
            self.tree.insert(
                "",
                tk.END,
                values=(idx, student['name'], student['phone'], status_text, payment_status),
                tags=(combined_tag, str(student['row']))
            )
        
        # 태그 스타일
        self.tree.tag_configure("checkedin_paid", background="#E8F5E9")
        self.tree.tag_configure("checkedout_paid", background="#FFEBEE")
        self.tree.tag_configure("checkedin_unpaid", background="#FFF9C4")
        self.tree.tag_configure("checkedout_unpaid", background="#FFCDD2")
        
        # 상태 표시줄 업데이트
        self.status_label.config(
            text=f"마지막 업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 총 {len(students)}명"
        )
    
    def get_selected_student(self):
        """선택된 학생 정보 가져오기"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("경고", "학생을 선택해주세요.")
            return None
        
        item = self.tree.item(selection[0])
        values = item['values']
        tags = item['tags']
        
        return {
            'name': values[1],
            'phone': values[2],
            'status_text': values[3],
            'row': int(tags[1]),
            'status': 1 if "checkedin" in tags else 0
        }
    
    def update_status(self, row, new_status):
        """학생 상태 업데이트"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            status_cell = f"{self.config['status_column']}{row}"
            ws[status_cell].value = new_status
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            messagebox.showerror("오류", f"상태 업데이트 오류: {e}")
            return False
    
    def checkin(self):
        """등원 처리"""
        student = self.get_selected_student()
        if not student:
            return
        
        if student['status'] == 1:
            messagebox.showinfo("안내", f"{student['name']}님은 이미 등원중입니다.")
            return
        
        if self.update_status(student['row'], 1):
            timestamp = datetime.now().strftime('%H:%M:%S')
            message = f"[{timestamp}] {student['name']}님 등원 처리 완료\n\n"
            message += f'"{student["name"]}"님이 "{self.config.get("academy_name", "OO학원")}"에 등원하였습니다.'
            messagebox.showinfo("등원 완료", message)
            self.load_students()
    
    def checkout(self):
        """하원 처리"""
        student = self.get_selected_student()
        if not student:
            return
        
        if student['status'] == 0:
            messagebox.showinfo("안내", f"{student['name']}님은 이미 하원 상태입니다.")
            return
        
        if self.update_status(student['row'], 0):
            timestamp = datetime.now().strftime('%H:%M:%S')
            message = f"[{timestamp}] {student['name']}님 하원 처리 완료\n\n"
            message += f'"{student["name"]}"님이 "{self.config.get("academy_name", "OO학원")}"에서 하원하였습니다.'
            messagebox.showinfo("하원 완료", message)
            self.load_students()
    
    def register_payment(self):
        """원비 납입 등록"""
        student = self.get_selected_student()
        if not student:
            return
        
        # 납입일 입력 다이얼로그
        from tkinter import simpledialog
        payment_date = simpledialog.askstring(
            "원비 납입 등록",
            f"{student['name']}님의 납입일을 입력하세요.\n(예: 2024-01-15 또는 01/15)",
            parent=self.root
        )
        
        if payment_date:
            if self.update_payment(student['row'], payment_date):
                messagebox.showinfo("완료", f"{student['name']}님 원비 납입 등록 완료\n납입일: {payment_date}")
                self.load_students()
        else:
            # 납입일 삭제 (미납으로 변경)
            confirm = messagebox.askyesno("확인", "납입 정보를 삭제하시겠습니까?")
            if confirm:
                if self.update_payment(student['row'], None):
                    messagebox.showinfo("완료", f"{student['name']}님 납입 정보 삭제 완료")
                    self.load_students()
    
    def update_payment(self, row, payment_date):
        """납입 정보 업데이트"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            payment_cell = f"{self.config['payment_column']}{row}"
            ws[payment_cell].value = payment_date
            
            wb.save(self.excel_file)
            wb.close()
            return True
            
        except Exception as e:
            messagebox.showerror("오류", f"납입 정보 업데이트 오류: {e}")
            return False
    
    def send_checkin_message(self):
        """등원 알림 수동 발송"""
        student = self.get_selected_student()
        if not student:
            return
        
        academy_name = self.config.get('academy_name', 'OO학원')
        message = f'"{student["name"]}"님이 "{academy_name}"에 등원하였습니다.'
        
        if not student['phone']:
            messagebox.showwarning("경고", f"{student['name']}님의 연락처가 없습니다.")
            return
        
        confirm = messagebox.askyesno(
            "등원 알림 발송", 
            f"{student['name']}님에게 등원 알림을 발송하시겠습니까?\n\n"
            f"수신: {student['phone']}\n"
            f"내용: {message}"
        )
        
        if confirm:
            from sms_sender import send_sms
            success = send_sms(student['phone'], message, student_name=student['name'])
            
            if success:
                timestamp = datetime.now().strftime('%H:%M:%S')
                messagebox.showinfo("발송 완료", f"[{timestamp}] {student['name']}님에게 등원 알림 발송 완료")
            else:
                messagebox.showerror("발송 실패", "메시지 발송에 실패했습니다.")
    
    def send_checkout_message(self):
        """하원 알림 수동 발송"""
        student = self.get_selected_student()
        if not student:
            return
        
        academy_name = self.config.get('academy_name', 'OO학원')
        message = f'"{student["name"]}"님이 "{academy_name}"에서 하원하였습니다.'
        
        if not student['phone']:
            messagebox.showwarning("경고", f"{student['name']}님의 연락처가 없습니다.")
            return
        
        confirm = messagebox.askyesno(
            "하원 알림 발송", 
            f"{student['name']}님에게 하원 알림을 발송하시겠습니까?\n\n"
            f"수신: {student['phone']}\n"
            f"내용: {message}"
        )
        
        if confirm:
            from sms_sender import send_sms
            success = send_sms(student['phone'], message, student_name=student['name'])
            
            if success:
                timestamp = datetime.now().strftime('%H:%M:%S')
                messagebox.showinfo("발송 완료", f"[{timestamp}] {student['name']}님에게 하원 알림 발송 완료")
            else:
                messagebox.showerror("발송 실패", "메시지 발송에 실패했습니다.")
    
    def send_payment_request(self):
        """납입 요청 문자 발송"""
        student = self.get_selected_student()
        if not student:
            return
        
        # 이미 납입한 학생인지 확인
        if student.get('payment_date'):
            confirm = messagebox.askyesno(
                "확인", 
                f"{student['name']}님은 이미 납입 완료 상태입니다.\n"
                f"(납입일: {student['payment_date']})\n\n"
                f"그래도 납입 요청 문자를 발송하시겠습니까?"
            )
            if not confirm:
                return
        
        academy_name = self.config.get('academy_name', 'OO학원')
        
        # 납입 요청 메시지 (사용자 정의 가능)
        from tkinter import simpledialog
        default_message = f'안녕하세요, {academy_name}입니다.\n{student["name"]}님의 이번 달 원비 납입을 부탁드립니다.\n문의사항은 학원으로 연락 주세요.'
        
        message = simpledialog.askstring(
            "납입 요청 메시지",
            f"{student['name']}님에게 보낼 메시지를 입력하세요:\n(수정 가능)",
            initialvalue=default_message,
            parent=self.root
        )
        
        if not message:
            return
        
        if not student['phone']:
            messagebox.showwarning("경고", f"{student['name']}님의 연락처가 없습니다.")
            return
        
        confirm = messagebox.askyesno(
            "납입 요청 발송", 
            f"{student['name']}님에게 납입 요청을 발송하시겠습니까?\n\n"
            f"수신: {student['phone']}\n"
            f"내용: {message}"
        )
        
        if confirm:
            from sms_sender import send_sms
            success = send_sms(student['phone'], message, student_name=student['name'])
            
            if success:
                timestamp = datetime.now().strftime('%H:%M:%S')
                messagebox.showinfo("발송 완료", f"[{timestamp}] {student['name']}님에게 납입 요청 발송 완료")
            else:
                messagebox.showerror("발송 실패", "메시지 발송에 실패했습니다.")
    
    def edit_phone(self):
        """연락처 수정"""
        student = self.get_selected_student()
        if not student:
            return
        
        from tkinter import simpledialog
        new_phone = simpledialog.askstring(
            "연락처 수정",
            f"{student['name']}님의 새 연락처를 입력하세요:\n\n"
            f"현재: {student['phone']}\n"
            f"형식: 010-1234-5678 또는 01012345678",
            initialvalue=student['phone'],
            parent=self.root
        )
        
        if new_phone is None:  # 취소
            return
        
        if not new_phone.strip():
            messagebox.showwarning("경고", "연락처를 입력해주세요.")
            return
        
        # 연락처 업데이트
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            phone_cell = f"{self.config['phone_column']}{student['row']}"
            ws[phone_cell].value = new_phone
            
            wb.save(self.excel_file)
            wb.close()
            
            messagebox.showinfo("완료", f"{student['name']}님 연락처 수정 완료\n새 연락처: {new_phone}")
            self.load_students()
            
        except Exception as e:
            messagebox.showerror("오류", f"연락처 수정 오류: {e}")
    
    def add_student(self):
        """신규 학생 등록"""
        from tkinter import simpledialog
        
        # 학생 이름 입력
        name = simpledialog.askstring(
            "학생 등록 - 1/3",
            "학생 이름을 입력하세요:",
            parent=self.root
        )
        
        if not name or not name.strip():
            return
        
        # 연락처 입력
        phone = simpledialog.askstring(
            "학생 등록 - 2/3",
            f"{name}님의 연락처를 입력하세요:\n(예: 010-1234-5678)",
            parent=self.root
        )
        
        if not phone or not phone.strip():
            messagebox.showwarning("경고", "연락처를 입력해주세요.")
            return
        
        # 납입일 입력 (선택사항)
        payment_date = simpledialog.askstring(
            "학생 등록 - 3/3",
            f"{name}님의 원비 납입일을 입력하세요:\n(선택사항, 빈칸이면 미납)\n(예: 2024-01-15)",
            parent=self.root
        )
        
        # 엑셀에 추가
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # 마지막 행 찾기
            last_row = self.config['start_row']
            while ws[f"{self.config['name_column']}{last_row}"].value:
                last_row += 1
            
            # 새 학생 정보 입력
            ws[f"{self.config['name_column']}{last_row}"].value = name.strip()
            ws[f"{self.config['phone_column']}{last_row}"].value = phone.strip()
            ws[f"{self.config['status_column']}{last_row}"].value = 0  # 하원 상태
            if payment_date and payment_date.strip():
                ws[f"{self.config['payment_column']}{last_row}"].value = payment_date.strip()
            
            wb.save(self.excel_file)
            wb.close()
            
            payment_info = f"\n납입일: {payment_date}" if payment_date and payment_date.strip() else "\n납입: 미납"
            messagebox.showinfo("완료", f"학생 등록 완료!\n\n이름: {name}\n연락처: {phone}{payment_info}")
            self.load_students()
            
        except Exception as e:
            messagebox.showerror("오류", f"학생 등록 오류: {e}")
    
    def delete_student(self):
        """학생 삭제"""
        student = self.get_selected_student()
        if not student:
            return
        
        # 확인
        confirm = messagebox.askyesno(
            "학생 삭제",
            f"{student['name']}님을 삭제하시겠습니까?\n\n"
            f"연락처: {student['phone']}\n"
            f"상태: {'등원중' if student['status'] == 1 else '하원'}\n\n"
            f"⚠️ 이 작업은 되돌릴 수 없습니다!"
        )
        
        if not confirm:
            return
        
        # 엑셀에서 삭제
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # 해당 행 삭제
            ws.delete_rows(student['row'], 1)
            
            wb.save(self.excel_file)
            wb.close()
            
            messagebox.showinfo("완료", f"{student['name']}님이 삭제되었습니다.")
            self.load_students()
            
        except Exception as e:
            messagebox.showerror("오류", f"학생 삭제 오류: {e}")
    
    def auto_refresh(self):
        """자동 새로고침 (3초마다)"""
        self.load_students()
        self.root.after(3000, self.auto_refresh)
    
    def run(self):
        """GUI 실행"""
        self.root.mainloop()

if __name__ == "__main__":
    try:
        app = AttendanceGUI()
        app.run()
    except Exception as e:
        messagebox.showerror("오류", f"프로그램 실행 오류: {e}")

