import openpyxl
import time
import json
import os
from datetime import datetime
from sms_sender import send_sms

class AttendanceMonitor:
    def __init__(self, excel_file, config_file='config.json'):
        """
        학원 등원/하원 모니터링 클래스
        
        Args:
            excel_file: 엑셀 파일 경로
            config_file: 설정 파일 경로
        """
        self.excel_file = excel_file
        self.config_file = config_file
        self.previous_state = {}
        self.config = self.load_config()
        
    def load_config(self):
        """설정 파일 로드"""
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            # 기본 설정
            default_config = {
                "academy_name": "OO학원",
                "check_interval": 5,  # 초 단위
                "name_column": "A",  # 이름 열
                "phone_column": "B",  # 연락처 열
                "status_column": "C",  # 등원/하원 상태 열
                "start_row": 2  # 데이터 시작 행
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, ensure_ascii=False, indent=2)
            return default_config
    
    def read_excel(self):
        """엑셀 파일에서 학생 정보 읽기"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            students = {}
            row = self.config['start_row']
            
            while True:
                name_cell = f"{self.config['name_column']}{row}"
                phone_cell = f"{self.config['phone_column']}{row}"
                status_cell = f"{self.config['status_column']}{row}"
                
                name = ws[name_cell].value
                
                # 이름이 없으면 종료
                if not name:
                    break
                
                phone = ws[phone_cell].value
                status = ws[status_cell].value
                
                # 상태 값을 0 또는 1로 변환
                if status is None:
                    status = 0
                else:
                    try:
                        status = int(status)
                    except:
                        status = 0
                
                students[name] = {
                    'phone': str(phone) if phone else '',
                    'status': status,
                    'row': row
                }
                
                row += 1
            
            wb.close()
            return students
            
        except Exception as e:
            print(f"엑셀 파일 읽기 오류: {e}")
            return {}
    
    def check_status_change(self):
        """상태 변화 확인 및 알림 발송"""
        current_state = self.read_excel()
        
        # 첫 실행 시 현재 상태만 저장
        if not self.previous_state:
            self.previous_state = current_state.copy()
            print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 모니터링 시작")
            print(f"현재 {len(current_state)}명의 학생 정보를 로드했습니다.")
            return
        
        # 상태 변화 확인
        for name, current_info in current_state.items():
            if name in self.previous_state:
                prev_status = self.previous_state[name]['status']
                curr_status = current_info['status']
                
                # 상태가 변경된 경우
                if prev_status != curr_status:
                    self.handle_status_change(name, current_info, prev_status, curr_status)
        
        # 상태 업데이트
        self.previous_state = current_state.copy()
    
    def handle_status_change(self, name, info, prev_status, curr_status):
        """상태 변화 처리 및 문자 발송"""
        phone = info['phone']
        academy_name = self.config['academy_name']
        
        # 0 -> 1: 등원
        if prev_status == 0 and curr_status == 1:
            action = "등원"
            message = f'"{name}"님이 "{academy_name}"에 등원하였습니다.'
            
        # 1 -> 0: 하원
        elif prev_status == 1 and curr_status == 0:
            action = "하원"
            message = f'"{name}"님이 "{academy_name}"에서 하원하였습니다.'
        else:
            # 그 외의 경우는 무시
            return
        
        # 로그 출력
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[{timestamp}] {name} - {action} (연락처: {phone})")
        print(f"  메시지: {message}")
        
        # SMS/카카오톡 전송
        if phone and phone.strip():
            success = send_sms(phone, message, student_name=name)
            if success:
                print(f"  ✓ 메시지 전송 성공")
            else:
                print(f"  ✗ 메시지 전송 실패")
        else:
            print(f"  ✗ 연락처 없음 - 문자 미전송")
        
        print()
    
    def start_monitoring(self):
        """모니터링 시작"""
        print("=" * 60)
        print(f"학원 등원/하원 알림 시스템 시작")
        print(f"학원명: {self.config['academy_name']}")
        print(f"체크 주기: {self.config['check_interval']}초")
        print(f"엑셀 파일: {self.excel_file}")
        print("=" * 60)
        print()
        
        try:
            while True:
                self.check_status_change()
                time.sleep(self.config['check_interval'])
        except KeyboardInterrupt:
            print("\n모니터링을 종료합니다.")
        except Exception as e:
            print(f"오류 발생: {e}")

if __name__ == "__main__":
    # 엑셀 파일 경로
    excel_file = "202511_자동알림.xlsx"
    
    # 모니터 인스턴스 생성 및 시작
    monitor = AttendanceMonitor(excel_file)
    monitor.start_monitoring()

